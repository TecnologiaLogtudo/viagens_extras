import openpyxl
import xlrd
import os
from typing import Callable, Tuple
import re
import unicodedata

# !!! ATENÇÃO: Defina o índice da coluna "Remetente" aqui !!!
# O índice da coluna começa em 0. Se a coluna for 'A', o índice é 0; 'B' é 1; 'C' é 2, etc.
# Substitua -1 pelo número da coluna correta da planilha original.
# Com base em layouts comuns, estamos assumindo a coluna G (índice 6). Ajuste se necessário.
REMETENTE_COLUMN_INDEX = 11

# Regex para encontrar placas - ATUALIZADO
# Cobre:
# - ABC-1234 / ABC1234 / ABC 1234 (Antigas)
# - ABC-1D23 / ABC1D23 / ABC 1D23 (Mercosul)
# - MQTIE66 / MQT IE66 (XXXXXNN / XXX XXNN)
PLACA_REGEX = re.compile(
    # (Grupo 1: Formatos [A-Z]{3}[- ]?[A-Z0-9]{4})
    r'((?:[A-Z]{3}[ -]?(?:\d{4}|\d[A-Z]\d{2}))' + 
    # | (OU)
    r'|' + 
    # (Grupo 2: Formatos [A-Z]{3} ?[A-Z]{2}\d{2})
    r'(?:[A-Z]{3} ?[A-Z]{2}\d{2}))', 
    re.IGNORECASE # Ignora se é maiúscula ou minúscula
)


# ATUALIZAÇÃO: Regex para limpar padrões indesejados do "Nome"
# Horários (NN:NN, NNhNN, NNh), "h" sozinho, e textos ("VIAGEM EXTRA", "RASTREADOR")
NOME_CLEANUP_REGEX = re.compile(
    r'(\d{1,2}:\d{2}|\d{1,2}h\d{2}|\d{1,2}h(?!\d)|h(?!\w)|VIAGEM EXTRA|RASTREADOR)',
    re.IGNORECASE
)

# Regex para encontrar a data de pagamento (NN/NN/NN ou NN/NN/NNNN)
DATA_PAGAMENTO_REGEX = re.compile(r'(\d{2}/\d{2}/\d{2,4})')

def processar_planilha(filepath: str, log_callback: Callable) -> str | None:
    """
    Função principal da Fase 1.
    Lê uma planilha, processa os dados, e salva um novo arquivo .xlsx com duas abas:
    'Dados Processados' para linhas válidas e 'Contrato não realizado' para linhas com dados faltantes.
    """
    try:
        if REMETENTE_COLUMN_INDEX == -1:
            log_callback("[F1] ERRO: O índice da coluna 'Remetente' não foi definido.", "ERRO")
            log_callback("[F1] Edite o arquivo 'fase1_processamento.py' e defina a variável 'REMETENTE_COLUMN_INDEX' com o número da coluna correta.", "ERRO")
            return None
            
        # Define o nome do arquivo de saída.
        # ALTERAÇÃO: Em vez de salvar ao lado do arquivo de entrada,
        # vamos salvar na pasta onde o programa está rodando (a pasta do .exe).
        base_dir = os.getcwd() 
        filename = os.path.splitext(os.path.basename(filepath))[0]
        output_filename = os.path.join(base_dir, f"{filename}_processado.xlsx")

        log_callback("[F1] Iniciando processamento da planilha...", "INFO")

        wb_out = openpyxl.Workbook()
        
        # Aba para dados processados com sucesso
        ws_out = wb_out.active
        ws_out.title = "Dados Processados"

        headers = ["Nro cotação", "Categoria veículo", "Cidade", "UF", "Nome", "Placa", "Data pagamento", "Viagem extra", "Remetente", "Status"]
        ws_out.append(headers)

        # Aba para contratos não realizados
        ws_nao_realizado = wb_out.create_sheet("Contrato não realizado")
        ws_nao_realizado.sheet_properties.tabColor = "FF0000"  # Cor vermelha
        headers_nao_realizado = ["Nro cotação", "Categoria veículo", "Cidade", "UF", "Nome", "Placa", "Data pagamento", "Observação", "Status"]
        ws_nao_realizado.append(headers_nao_realizado)

        if filepath.endswith('.xlsx'):
            log_callback("[F1] Detectado formato .xlsx (openpyxl).", "DEBUG")
            _ler_com_openpyxl(filepath, ws_out, ws_nao_realizado, log_callback)
        elif filepath.endswith('.xls'):
            log_callback("[F1] Detectado formato .xls (xlrd).", "DEBUG")
            _ler_com_xlrd(filepath, ws_out, ws_nao_realizado, log_callback)
        else:
            log_callback(f"[F1] Extensão de arquivo não suportada: {filepath}", "ERRO")
            return None

        wb_out.save(output_filename)
        log_callback(f"[F1] Nova planilha salva em: {output_filename}", "SUCESSO")
        return output_filename

    except PermissionError:
        # Tratamento específico para erro de permissão (arquivo aberto)
        output_basename = os.path.basename(output_filename) if 'output_filename' in locals() else "o arquivo de saída"
        log_callback(f"[F1] ERRO DE PERMISSÃO AO SALVAR A PLANILHA.", "ERRO")
        log_callback(f"[F1] Verifique se a planilha '{output_basename}' já está aberta em seu computador. Se estiver, feche-a e tente novamente.", "ERRO")
        log_callback(f"[F1] A automação não pode continuar pois não consegue criar o arquivo de resultado.", "ERRO")
        return None
    except Exception as e:
        log_callback(f"[F1] Erro inesperado no processamento da planilha: {e}", "ERRO")
        import traceback
        log_callback(f"Traceback: {traceback.format_exc()}", "DEBUG")
        return None

def _ler_com_openpyxl(filepath: str, ws_out: openpyxl.worksheet.worksheet.Worksheet, ws_nao_realizado: openpyxl.worksheet.worksheet.Worksheet, log_callback: Callable):
    wb_in = None
    try:
        wb_in = openpyxl.load_workbook(filepath, data_only=True)
        ws_in = wb_in.active
        
        if hasattr(ws_in, '_images') and ws_in._images:
            ws_in._images = []
            log_callback("[F1] Imagens removidas da planilha de origem.", "DEBUG")

        log_callback(f"[F1] Lendo dados de {ws_in.max_row} linhas (openpyxl)...", "DEBUG")

        # Começa da linha 9 para pular o cabeçalho antigo que pode estar na linha 8
        for row_idx, row in enumerate(ws_in.iter_rows(min_row=9, values_only=True), start=9):
            try:
                # Col B (Nro cotação) -> Índice 1
                # Col K (Categoria)   -> Índice 10
                # Col M (Cidade/UF)   -> Índice 12
                # Col W (Nome/Placa)  -> Índice 22
                
                # ATUALIZAÇÃO: Pula linha se Nro Cotação (col B) estiver vazio
                if row[1] is None or str(row[1]).strip() == "":
                    log_callback(f"[F1] Pulando linha {row_idx} (Nro cotação vazio).", "DEBUG")
                    continue

                nro_cotacao = row[1]
                categoria = row[10]
                cidade_uf_raw = row[12]
                nome_placa_raw = row[22]
                remetente_raw = row[REMETENTE_COLUMN_INDEX] if REMETENTE_COLUMN_INDEX < len(row) else None
                
                cidade, uf = _processar_cidade_uf(cidade_uf_raw, row_idx, log_callback)
                nome, placa, data_pagamento, viagem_extra = _processar_nome_placa(nome_placa_raw, row_idx, log_callback)
                remetente = _processar_remetente(remetente_raw)

                # Validação dos dados
                erros = []
                if nome == "NOME NÃO ENCONTRADO":
                    erros.append("Nome")
                if placa == "PLACA NÃO ENCONTRADA":
                    erros.append("Placa")
                if data_pagamento == "DATA NÃO ENCONTRADA":
                    erros.append("Data pagamento")

                linha_base = [nro_cotacao, categoria, cidade, uf, nome, placa, data_pagamento]

                if erros:
                    observacao = f"Dados faltantes: {', '.join(erros)}"
                    ws_nao_realizado.append(linha_base + [observacao, "Falha na Validação"])
                    log_callback(f"[F1] Linha {row_idx}: Movida para 'Contrato não realizado'. Motivo: {observacao}", "AVISO")
                else:
                    ws_out.append(linha_base + [viagem_extra, remetente, "Pendente"])

            except IndexError:
                log_callback(f"[F1] Erro ao ler colunas na linha {row_idx}. A linha pode ser mais curta que o esperado.", "AVISO")
            except Exception as e_row:
                 log_callback(f"[F1] Erro processando linha {row_idx}: {e_row}", "AVISO")

        log_callback("[F1] Leitura (openpyxl) concluída.", "DEBUG")

    finally:
        if wb_in:
            wb_in.close()

def _get_merged_cell_value(sheet: xlrd.sheet.Sheet, row_idx: int, col_idx: int) -> any:
    """
    Helper para xlrd. Encontra o valor de uma célula, mesmo se ela estiver mesclada.
    """
    for (rlo, rhi, clo, chi) in sheet.merged_cells:
        if rlo <= row_idx < rhi and clo <= col_idx < chi:
            return sheet.cell_value(rlo, clo)
    return sheet.cell_value(row_idx, col_idx)

def _ler_com_xlrd(filepath: str, ws_out: openpyxl.worksheet.worksheet.Worksheet, ws_nao_realizado: openpyxl.worksheet.worksheet.Worksheet, log_callback: Callable):
    """
    Lógica de leitura para arquivos .xls.
    """
    wb_in = None
    try:
        wb_in = xlrd.open_workbook(filepath, formatting_info=False)
        sheet = wb_in.sheet_by_index(0)

        log_callback(f"[F1] Lendo dados de {sheet.nrows} linhas (xlrd)...", "DEBUG")

        # Começa do índice 8 (linha 9 no Excel) para pular o cabeçalho antigo
        for row_idx in range(8, sheet.nrows):
            try:
                # Col B (Nro cotação) -> Índice 1
                # Col K (Categoria)   -> Índice 10
                # Col M (Cidade/UF)   -> Índice 12
                # Col W (Nome/Placa)  -> Índice 22

                nro_cotacao = _get_merged_cell_value(sheet, row_idx, 1)

                # ATUALIZAÇÃO: Pula linha se Nro Cotação (col B) estiver vazio
                if nro_cotacao is None or str(nro_cotacao).strip() == "":
                    log_callback(f"[F1] Pulando linha {row_idx + 1} (Nro cotação vazio).", "DEBUG")
                    continue
                
                categoria = _get_merged_cell_value(sheet, row_idx, 10)
                cidade_uf_raw = _get_merged_cell_value(sheet, row_idx, 12)
                nome_placa_raw = _get_merged_cell_value(sheet, row_idx, 22)
                remetente_raw = _get_merged_cell_value(sheet, row_idx, REMETENTE_COLUMN_INDEX) if REMETENTE_COLUMN_INDEX < sheet.ncols else None
                
                cidade, uf = _processar_cidade_uf(cidade_uf_raw, row_idx + 1, log_callback)
                nome, placa, data_pagamento, viagem_extra = _processar_nome_placa(nome_placa_raw, row_idx + 1, log_callback)
                remetente = _processar_remetente(remetente_raw)

                # Validação dos dados
                erros = []
                if nome == "NOME NÃO ENCONTRADO":
                    erros.append("Nome")
                if placa == "PLACA NÃO ENCONTRADA":
                    erros.append("Placa")
                if data_pagamento == "DATA NÃO ENCONTRADA":
                    erros.append("Data pagamento")

                linha_base = [nro_cotacao, categoria, cidade, uf, nome, placa, data_pagamento]

                if erros:
                    observacao = f"Dados faltantes: {', '.join(erros)}"
                    ws_nao_realizado.append(linha_base + [observacao, "Falha na Validação"])
                    log_callback(f"[F1] Linha {row_idx + 1}: Movida para 'Contrato não realizado'. Motivo: {observacao}", "AVISO")
                else:
                    ws_out.append(linha_base + [viagem_extra, remetente, "Pendente"])

            except IndexError:
                log_callback(f"[F1] Erro ao ler colunas na linha {row_idx + 1}. A linha pode ser mais curta que o esperado.", "AVISO")
            except Exception as e_row:
                 log_callback(f"[F1] Erro processando linha {row_idx + 1}: {e_row}", "AVISO")

        log_callback("[F1] Leitura (xlrd) concluída.", "DEBUG")

    finally:
        pass # xlrd não tem .close()

# --- NOVAS FUNÇÕES HELPER (FASE 1) ---

def _processar_remetente(raw_string: any) -> str:
    """Extrai o número inicial da string do remetente."""
    if not raw_string or not isinstance(raw_string, str):
        return "N/A"
    
    # Remove espaços em branco no início e no fim
    cleaned_string = str(raw_string).strip()
    
    # Pega a primeira parte antes do " - "
    parts = cleaned_string.split(' - ')
    
    if parts:
        # Retorna a primeira parte, que deve ser a numeração
        return parts[0].strip()
    
    return "N/A"

def _remover_acentos(texto: str) -> str:
    """Remove acentos de uma string."""
    nfkd_form = unicodedata.normalize('NFKD', texto)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])

def _padronizar_cidade(cidade: str) -> str:
    """Padroniza nomes de cidades para correspondências específicas."""
    # Mapeamento com chaves normalizadas (maiúsculas, sem acentos)
    mapeamento_cidades = {
        "JOAO PESSOA": "J. Pessoa",
        "BAYEUX": "J. Pessoa",
        "CAMACARI": "Salvador",
        "SIMOES FILHO": "Salvador",
        "SERRA": "Vitória",
        "CARIACICA": "Vitória",
        "VILA VELHA": "Vitória",
    }
    # Normaliza a entrada para maiúsculas e remove acentos
    cidade_normalizada = _remover_acentos(cidade.upper())
    
    # Retorna o valor do mapa ou a cidade original se não houver correspondência
    return mapeamento_cidades.get(cidade_normalizada, cidade)

def _processar_cidade_uf(raw_string: any, line_num: int, log_callback: Callable) -> Tuple[str, str]:
    """Separa 'Cidade/UF' em duas colunas e padroniza o nome da cidade."""
    cidade = "CIDADE NÃO ENCONTRADA"
    uf = "UF NÃO ENCONTRADA"

    if not isinstance(raw_string, str) or "/" not in raw_string:
        log_callback(f"[F1] Linha {line_num}: 'Cidade/UF' não contém '/' ou não é string. (Valor: '{raw_string}').", "AVISO")
        cidade = str(raw_string).strip()
    else:
        parts = raw_string.split('/')
        cidade = parts[0].strip()
        uf = parts[1].strip()
        if len(parts) > 2:
            log_callback(f"[F1] Linha {line_num}: 'Cidade/UF' contém múltiplas '/'. Usando a primeira como Cidade. (Valor: '{raw_string}').", "AVISO")
            uf = "/".join(parts[1:]).strip()
    
    # Padroniza a cidade
    cidade_padronizada = _padronizar_cidade(cidade)
    if cidade_padronizada != cidade:
        log_callback(f"[F1] Linha {line_num}: Cidade '{cidade}' padronizada para '{cidade_padronizada}'.", "INFO")
        
    return cidade_padronizada, uf

def _processar_nome_placa(raw_string: any, line_num: int, log_callback: Callable) -> Tuple[str, str, str, str]:
    """Separa 'Nome/Placa', encontra a data de pagamento, verifica por 'Viagem extra' e retorna os quatro valores."""
    raw_str = str(raw_string).strip()
    placa = "PLACA NÃO ENCONTRADA"
    # Adiciona a variável para a data de pagamento
    data_pagamento = "DATA NÃO ENCONTRADA"
    nome_bruto = ""

    # --- ETAPA 1: Verificar "Viagem Extra" ---
    viagem_extra = "Sim" if re.search(r'viagem extra', raw_str, re.IGNORECASE) else "Não"

    # --- ETAPA 2: Encontrar a Placa ---
    match_placa = PLACA_REGEX.search(raw_str)
    
    if match_placa:
        # Placa encontrada
        placa_encontrada = match_placa.group(1).upper().replace(" ", "").replace("-", "")
        placa = placa_encontrada
        
        # Remove a placa da string original para processar o resto
        nome_bruto = PLACA_REGEX.sub('', raw_str).strip()
    else:
        # Placa não encontrada
        log_callback(f"[F1] Linha {line_num}: Placa não encontrada. (Valor: '{raw_str}')", "AVISO")
        # Assume que a string inteira é o nome (se não estiver vazia)
        if raw_str:
            nome_bruto = raw_str

    # --- ETAPA 3: Encontrar a Data de Pagamento ---
    # Procura a data no que restou da string
    match_data = DATA_PAGAMENTO_REGEX.search(nome_bruto)
    if match_data:
        # Data encontrada
        data_pagamento = match_data.group(1)
        # Remove a data da string para que não seja confundida com o nome
        nome_bruto = DATA_PAGAMENTO_REGEX.sub('', nome_bruto).strip()
    
    # --- ETAPA 4: Limpeza do Nome ---
    if nome_bruto:
        # 1. Remove prefixos de tempo como "as" ou "às" que precedem uma hora.
        nome_limpo = re.sub(r'(as|às)\s+(?=\d{1,2}:\d{2}|\d{1,2}h\d{2}|\d{1,2}h(?!\d)|h(?!\w))', '', nome_bruto, flags=re.IGNORECASE)
        # 2. Remove os padrões de tempo restantes e outras palavras-chave.
        nome_limpo = NOME_CLEANUP_REGEX.sub('', nome_limpo)
        # 3. Remove hífens
        nome_limpo = nome_limpo.replace('-', '')
        # 4. Remove espaços duplos/múltiplos
        nome_limpo = re.sub(r'\s+', ' ', nome_limpo).strip()
        
        # 5. Verifica se o nome limpo é válido
        if not nome_limpo:
            nome = "NOME NÃO ENCONTRADO"
            log_callback(f"[F1] Linha {line_num}: Nome não encontrado após limpeza. (Bruto: '{nome_bruto}')", "AVISO")
        # 6. Verifica se contém números ou códigos HTML
        elif re.search(r'\d', nome_limpo) or '&#' in nome_limpo:
            nome = "NOME NÃO ENCONTRADO"
            log_callback(f"[F1] Linha {line_num}: Nome inválido (contém números/códigos) após limpeza. (Nome limpo: '{nome_limpo}')", "AVISO")
        else:
            # 7. Nome é válido
            nome = nome_limpo
    else:
        # Se o nome_bruto já estava vazio
        nome = "NOME NÃO ENCONTRADO"
        if placa == "PLACA NÃO ENCONTRADA":
             log_callback(f"[F1] Linha {line_num}: Nome, Placa e Data não encontrados (Valor vazio).", "AVISO")
        else:
            # Isso acontece se a string original era SÓ a placa (e talvez a data)
            log_callback(f"[F1] Linha {line_num}: Placa '{placa}' encontrada, mas Nome não. (Valor: '{raw_str}')", "AVISO")
            
    # Retorna os quatro valores
    return nome, placa, data_pagamento, viagem_extra

if __name__ == "__main__":
    # Teste rápido da função de processamento (sem interface gráfica)
    def log_callback(message: str, level: str = "INFO"):
        print(f"[{level}] {message}")

    # Substitua pelo caminho do arquivo de teste
    test_filepath = "E:\\Downloads\\relatorio (48).xls"
    processar_planilha(test_filepath, log_callback)